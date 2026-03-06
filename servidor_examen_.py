from flask import Flask, request, jsonify, session, redirect, url_for, send_file
from flask_cors import CORS
import json, os, socket, base64, io
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'tecmilenio_examenes_2026'
CORS(app, supports_credentials=True)

# ── Carpetas ──────────────────────────────────────────────────
for d in ['respuestas', 'imagenes']:
    os.makedirs(d, exist_ok=True)

# ── Configuración por defecto ─────────────────────────────────

# Estado de edicion - bloquea el examen mientras el docente edita
import threading
_estado_edicion = {'editando': False, 'version': 0}
_estado_lock = threading.Lock()

@app.route('/estado_examen')
def estado_examen():
    return jsonify(_estado_edicion)

CONFIG_FILE  = 'config_examen.json'
ADMIN_PASS   = 'admin123'   # ← Cambia esta contraseña

CONFIG_DEFAULT = {
    "titulo": "Examen de Pensamiento Crítico",
    "materia": "Gestión de Redes",
    "tiempo_minutos": 0,   # 0 = sin límite
    "preguntas": [
        {
            "id": 1,
            "tipo": "opcion_multiple",
            "texto": "¿Cuál de las siguientes afirmaciones refleja mejor la idea de que el trabajo puede ser creativo o alienante según su contexto?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "El trabajo siempre es productivo si genera dinero"},
                {"id": "b", "texto": "El trabajo tiene valor cuando quien lo realiza comprende y controla el proceso"},
                {"id": "c", "texto": "Todo trabajo es igualmente valioso sin importar las condiciones"},
                {"id": "d", "texto": "El trabajo solo es alienante en fábricas"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 2,
            "tipo": "opcion_multiple",
            "texto": "En una sociedad donde muchas personas realizan tareas repetitivas sin entender su propósito final, ¿qué consecuencia es más probable?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Mayor eficiencia económica en todos los sectores"},
                {"id": "b", "texto": "Pérdida de sentido y desconexión de las personas con su trabajo"},
                {"id": "c", "texto": "Aumento automático de la satisfacción laboral"},
                {"id": "d", "texto": "Eliminación de todas las jerarquías organizacionales"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 3,
            "tipo": "opcion_multiple",
            "texto": "¿Qué característica distingue a un \"trabajo sin sentido\" de uno significativo?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "El salario que se recibe"},
                {"id": "b", "texto": "Si la persona misma cree que su trabajo contribuye a algo valioso"},
                {"id": "c", "texto": "La cantidad de horas trabajadas"},
                {"id": "d", "texto": "El nivel educativo requerido"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 4,
            "tipo": "opcion_multiple",
            "texto": "Si en una empresa las decisiones las toman solo los altos ejecutivos sin consultar a quienes realizan el trabajo directo, ¿qué tipo de relación laboral se establece?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Una relación de colaboración horizontal"},
                {"id": "b", "texto": "Una relación jerárquica donde los trabajadores están alienados del proceso de decisión"},
                {"id": "c", "texto": "Una organización perfectamente democrática"},
                {"id": "d", "texto": "Un modelo donde todos tienen igual poder"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 5,
            "tipo": "opcion_multiple",
            "texto": "¿Cuál de estos ejemplos ilustra mejor un trabajo que genera \"valor social\" más allá del valor económico?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Un trader que genera millones en la bolsa de valores"},
                {"id": "b", "texto": "Una enfermera que cuida pacientes en una comunidad rural"},
                {"id": "c", "texto": "Un consultor que optimiza despidos corporativos"},
                {"id": "d", "texto": "Un especulador inmobiliario que infla precios"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 6,
            "tipo": "opcion_multiple",
            "texto": "En antropología, cuando estudiamos diferentes culturas, ¿qué enfoque es más ético y científico?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Juzgar todas las prácticas según los estándares de nuestra propia cultura"},
                {"id": "b", "texto": "Intentar comprender cada práctica dentro de su propio contexto cultural"},
                {"id": "c", "texto": "Asumir que la cultura occidental es superior a todas las demás"},
                {"id": "d", "texto": "Ignorar las diferencias culturales y tratarlas como irrelevantes"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 7,
            "tipo": "opcion_multiple",
            "texto": "¿Qué demuestra el hecho de que diferentes sociedades organizan el intercambio de bienes de formas muy distintas (trueque, regalo, mercado, etc.)?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Que solo el capitalismo moderno es eficiente"},
                {"id": "b", "texto": "Que las formas de organización económica son construcciones culturales, no leyes naturales"},
                {"id": "c", "texto": "Que las sociedades primitivas no entienden la economía"},
                {"id": "d", "texto": "Que el dinero siempre ha existido en todas las culturas"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 8,
            "tipo": "opcion_multiple",
            "texto": "Si una persona siente que su trabajo es inútil pero lo mantiene por necesidad económica, esto refleja principalmente:",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Falta de ética laboral de esa persona"},
                {"id": "b", "texto": "Una contradicción entre supervivencia económica y realización personal"},
                {"id": "c", "texto": "Que todos los trabajos son igualmente satisfactorios"},
                {"id": "d", "texto": "Que el sistema laboral actual es perfecto"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 9,
            "tipo": "opcion_multiple",
            "texto": "¿Qué revela el concepto de \"deuda\" cuando se estudia históricamente en diferentes culturas?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Que siempre ha sido una relación puramente económica"},
                {"id": "b", "texto": "Que ha tenido significados morales, sociales y políticos variables según el contexto"},
                {"id": "c", "texto": "Que los bancos son una invención reciente sin precedentes históricos"},
                {"id": "d", "texto": "Que todas las culturas han tratado la deuda exactamente igual"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        },
        {
            "id": 10,
            "tipo": "opcion_multiple",
            "texto": "En el contexto de las redes informáticas, ¿qué principio organizativo se asemeja más a las formas cooperativas de organización social?",
            "imagen": "",
            "opciones": [
                {"id": "a", "texto": "Un sistema completamente centralizado donde un servidor controla todo"},
                {"id": "b", "texto": "Una red distribuida donde múltiples nodos comparten responsabilidades"},
                {"id": "c", "texto": "Un modelo donde solo una computadora puede tomar decisiones"},
                {"id": "d", "texto": "Una estructura jerárquica rígida sin redundancia"}
            ],
            "respuesta_correcta": "b",
            "puntos": 1
        }
    ]
}

def cargar_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return CONFIG_DEFAULT.copy()

def guardar_config(cfg):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

def calcular_calificacion(respuestas_alumno, preguntas):
    puntos_total  = sum(p.get('puntos', 1) for p in preguntas if p['tipo'] != 'abierta')
    puntos_obtenidos = 0
    correctas = 0
    for p in preguntas:
        if p['tipo'] == 'abierta':
            continue
        resp = respuestas_alumno.get(f"pregunta_{p['id']}", "")
        correcta = p.get('respuesta_correcta', '')
        if p['tipo'] == 'seleccion_multiple':
            dadas    = set(resp) if isinstance(resp, list) else set()
            esperadas = set(correcta) if isinstance(correcta, list) else set()
            if dadas == esperadas:
                puntos_obtenidos += p.get('puntos', 1)
                correctas += 1
        else:
            if str(resp).strip().lower() == str(correcta).strip().lower():
                puntos_obtenidos += p.get('puntos', 1)
                correctas += 1
    calificacion = round((puntos_obtenidos / puntos_total) * 10, 1) if puntos_total > 0 else 0
    return correctas, puntos_total, calificacion

def get_ips():
    ips = []
    try:
        hostname = socket.gethostname()
        raw = socket.getaddrinfo(hostname, None)
        ips = list(set([i[4][0] for i in raw if ':' not in i[4][0] and i[4][0] != '127.0.0.1']))
    except:
        pass
    return ips or ['(revisa ipconfig)']

# ══════════════════════════════════════════════════════════════
#  RUTAS PÚBLICAS
# ══════════════════════════════════════════════════════════════

@app.route('/ping')
def ping():
    return jsonify({'status': 'ok', 'servidor': 'Tecmilenio Examenes'})

@app.route('/config_examen')
def config_examen():
    """El HTML carga el examen desde aquí"""
    cfg = cargar_config()
    # No mandar respuestas correctas al alumno
    safe = {k: v for k, v in cfg.items() if k != 'preguntas'}
    safe['preguntas'] = []
    for p in cfg['preguntas']:
        sp = {k: v for k, v in p.items() if k not in ('respuesta_correcta',)}
        safe['preguntas'].append(sp)
    return jsonify(safe)

@app.route('/guardar', methods=['POST'])
def guardar_respuestas():
    try:
        datos = request.json
        cfg   = cargar_config()
        ts    = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
        nombre = datos.get('nombre', 'anonimo').replace(' ', '_')
        correctas, total, calificacion = calcular_calificacion(
            datos.get('respuestas', {}), cfg['preguntas'])
        datos['calificacion'] = calificacion
        datos['correctas']    = correctas
        datos['total_puntos'] = total
        fname = f'respuestas/examen_{nombre}_{ts}.json'
        with open(fname, 'w', encoding='utf-8') as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
        print(f'[OK] {nombre} | Cal: {calificacion} ({correctas}/{total})')
        return jsonify({'status': 'ok', 'mensaje': 'Examen enviado correctamente',
                        'calificacion': calificacion, 'correctas': correctas, 'total': total})
    except Exception as e:
        print(f'[ERROR] {e}')
        return jsonify({'status': 'error', 'mensaje': str(e)}), 500

@app.route('/respuestas_correctas')
def respuestas_correctas():
    """Manda las respuestas correctas al alumno después de enviar"""
    cfg = cargar_config()
    correctas = {}
    for p in cfg['preguntas']:
        if p['tipo'] != 'abierta':
            correctas['pregunta_' + str(p['id'])] = p.get('respuesta_correcta', '')
    return jsonify(correctas)

@app.route('/imagen/<nombre>')
def servir_imagen(nombre):
    path = os.path.join('imagenes', nombre)
    if os.path.exists(path):
        return send_file(path)
    return '', 404

# ══════════════════════════════════════════════════════════════
#  AUTENTICACIÓN ADMIN
# ══════════════════════════════════════════════════════════════

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    error = ''
    if request.method == 'POST':
        is_json = 'application/json' in (request.content_type or '')
        pwd = (request.get_json(silent=True) or {}).get('password', '') if is_json else request.form.get('password', '')
        if pwd == ADMIN_PASS:
            session['admin'] = True
            return jsonify({'ok': True}) if is_json else redirect('/admin')
        error = 'Contraseña incorrecta'
        if is_json:
            return jsonify({'ok': False, 'error': error}), 401
    return f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Admin | Tecmilenio</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:linear-gradient(135deg,#00A859,#007A3D);min-height:100vh;display:flex;align-items:center;justify-content:center}}
.card{{background:white;border-radius:20px;padding:50px 40px;width:380px;box-shadow:0 20px 60px rgba(0,0,0,.3);text-align:center}}
img{{width:180px;margin-bottom:25px}}
h2{{color:#007A3D;margin-bottom:25px;font-size:22px}}
input{{width:100%;padding:14px;border:2px solid #ddd;border-radius:10px;font-size:16px;margin-bottom:15px;outline:none}}
input:focus{{border-color:#00A859}}
button{{width:100%;padding:15px;background:linear-gradient(135deg,#00A859,#007A3D);color:white;border:none;border-radius:10px;font-size:17px;font-weight:700;cursor:pointer}}
button:hover{{opacity:.9}}
.err{{color:#dc3545;margin-bottom:15px;font-size:14px}}
</style></head>
<body><div class="card">
<img src="https://upload.wikimedia.org/wikipedia/commons/6/65/Universidad_Tecmilenio.png" alt="Tecmilenio" style="width:200px;margin-bottom:20px">
<h2>Panel Docente</h2>
{"<p class='err'>"+error+"</p>" if error else ""}
<form method="POST">
<input type="password" name="password" placeholder="Contraseña" autofocus>
<button type="submit">Entrar</button>
</form>
</div></body></html>'''

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin'):
            return redirect('/admin/login')
        return f(*args, **kwargs)
    return decorated

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect('/admin/login')

# ══════════════════════════════════════════════════════════════
#  PANEL ADMIN
# ══════════════════════════════════════════════════════════════

@app.route('/admin')
@admin_required
def admin_home():
    cfg   = cargar_config()
    total = len(os.listdir('respuestas'))
    npregs = len(cfg['preguntas'])
    return f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Panel Docente | Tecmilenio</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f0f9f4;min-height:100vh}}
.nav{{background:linear-gradient(135deg,#00A859,#007A3D);padding:15px 30px;display:flex;align-items:center;justify-content:space-between;color:white}}
.nav img{{height:40px}}
.nav-links a{{color:white;text-decoration:none;margin-left:20px;font-weight:600;font-size:15px;padding:8px 16px;border-radius:20px;transition:.2s}}
.nav-links a:hover{{background:rgba(255,255,255,.2)}}
.container{{max-width:1100px;margin:30px auto;padding:0 20px}}
h1{{color:#007A3D;margin-bottom:25px;font-size:28px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:20px;margin-bottom:35px}}
.stat{{background:white;border-radius:15px;padding:25px;text-align:center;box-shadow:0 3px 15px rgba(0,0,0,.08);border-top:5px solid #00A859;cursor:pointer;transition:.2s;text-decoration:none;display:block;color:inherit}}
.stat:hover{{transform:translateY(-4px)}}
.stat .num{{font-size:52px;font-weight:900;color:#00A859}}
.stat .lbl{{color:#888;font-size:13px;text-transform:uppercase;letter-spacing:.5px;margin-top:5px}}
.btn{{display:inline-block;padding:14px 30px;border-radius:50px;font-size:16px;font-weight:700;text-decoration:none;transition:.2s;cursor:pointer;border:none}}
.btn-green{{background:linear-gradient(135deg,#00A859,#007A3D);color:white;box-shadow:0 5px 20px rgba(0,168,89,.3)}}
.btn-green:hover{{transform:translateY(-2px)}}
.btn-blue{{background:#0d6efd;color:white}}
.btn-red{{background:#dc3545;color:white}}
.actions{{display:flex;gap:15px;flex-wrap:wrap}}
</style></head>
<body>
<nav class="nav">
  <img src="https://upload.wikimedia.org/wikipedia/commons/6/65/Universidad_Tecmilenio.png" alt="Tecmilenio" style="height:36px;filter:brightness(0) invert(1)">
  <div class="nav-links">
    <a href="/">Inicio</a>
    <a href="/admin">Panel</a>
    <a href="/admin/examen">✏️ Editor de Examen</a>
    <a href="/ver">Resultados</a>
    <a href="/admin/descargar">📥 Descargar Excel</a>
    <a href="/admin/logout">Salir</a>
  </div>
</nav>
<div class="container">
  <h1>Panel Docente</h1>
  <div class="grid">
    <a href="/admin/examen" class="stat">
      <div class="num">{npregs}</div>
      <div class="lbl">Preguntas configuradas</div>
    </a>
    <a href="/ver" class="stat">
      <div class="num">{total}</div>
      <div class="lbl">Respuestas recibidas</div>
    </a>
    <a href="/admin/examen" class="stat" style="border-top-color:#0d6efd">
      <div class="num" style="color:#0d6efd;font-size:36px">✏️</div>
      <div class="lbl">Editar examen</div>
    </a>
    <a href="/admin/descargar" class="stat" style="border-top-color:#fd7e14">
      <div class="num" style="color:#fd7e14;font-size:36px">📥</div>
      <div class="lbl">Descargar resultados</div>
    </a>
  </div>
  <div class="actions">
    <a href="/admin/examen" class="btn btn-green">✏️ Editar Examen</a>
    <a href="/ver" class="btn btn-blue">Ver Resultados</a>
    <a href="/admin/descargar" class="btn" style="background:#fd7e14;color:white">📥 Descargar Excel</a>
  </div>
</div></body></html>'''

# ── Editor de examen ──────────────────────────────────────────

@app.route('/admin/examen', methods=['GET'])
@admin_required
def admin_examen():
    cfg = cargar_config()
    pregs_json = json.dumps(cfg['preguntas'], ensure_ascii=False)
    return f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Editor de Examen | Tecmilenio</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:#f0f9f4}}
.nav{{background:linear-gradient(135deg,#00A859,#007A3D);padding:15px 30px;display:flex;align-items:center;justify-content:space-between;color:white}}
.nav img{{height:40px}}
.nav-links a{{color:white;text-decoration:none;margin-left:20px;font-weight:600;padding:8px 16px;border-radius:20px}}
.nav-links a:hover{{background:rgba(255,255,255,.2)}}
.container{{max-width:1000px;margin:30px auto;padding:0 20px}}
h1{{color:#007A3D;margin-bottom:20px}}
.config-card{{background:white;border-radius:15px;padding:25px;margin-bottom:25px;box-shadow:0 3px 15px rgba(0,0,0,.08)}}
.config-card h3{{color:#007A3D;margin-bottom:15px;font-size:18px;border-bottom:2px solid #e8f5e9;padding-bottom:10px}}
.row{{display:grid;grid-template-columns:1fr 1fr;gap:15px}}
label{{display:block;font-weight:600;color:#444;margin-bottom:5px;font-size:14px}}
input[type=text],input[type=number],textarea,select{{width:100%;padding:10px 14px;border:2px solid #ddd;border-radius:8px;font-size:15px;outline:none;font-family:Arial}}
input:focus,textarea:focus,select:focus{{border-color:#00A859}}
.pregunta-card{{background:white;border-radius:15px;padding:25px;margin-bottom:20px;box-shadow:0 3px 15px rgba(0,0,0,.08);border-left:5px solid #00A859}}
.pregunta-card.abierta{{border-left-color:#0d6efd}}
.pregunta-card.seleccion_multiple{{border-left-color:#fd7e14}}
.pregunta-header{{display:flex;align-items:center;justify-content:space-between;margin-bottom:15px}}
.pregunta-num{{background:#00A859;color:white;width:32px;height:32px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:700;font-size:14px;flex-shrink:0}}
.opcion-row{{display:flex;align-items:center;gap:10px;margin-bottom:8px}}
.opcion-row input[type=text]{{flex:1}}
.opcion-id{{background:#e8f5e9;color:#007A3D;font-weight:700;padding:6px 10px;border-radius:6px;font-size:13px;min-width:28px;text-align:center}}
.btn{{padding:10px 20px;border-radius:50px;font-size:14px;font-weight:700;cursor:pointer;border:none;transition:.2s}}
.btn-green{{background:linear-gradient(135deg,#00A859,#007A3D);color:white}}
.btn-green:hover{{opacity:.9}}
.btn-red{{background:#dc3545;color:white}}
.btn-blue{{background:#0d6efd;color:white}}
.btn-gray{{background:#6c757d;color:white}}
.btn-orange{{background:#fd7e14;color:white}}
.btn-sm{{padding:6px 14px;font-size:12px}}
.tipo-badge{{padding:4px 12px;border-radius:20px;font-size:12px;font-weight:700}}
.tipo-opcion_multiple{{background:#e8f5e9;color:#007A3D}}
.tipo-abierta{{background:#dbeafe;color:#0d6efd}}
.tipo-seleccion_multiple{{background:#fff3cd;color:#856404}}
.correcta-check{{accent-color:#00A859;width:18px;height:18px;cursor:pointer}}
.img-preview{{max-width:200px;max-height:120px;border-radius:8px;margin-top:8px;display:none}}
.toolbar{{display:flex;gap:10px;margin-bottom:25px;flex-wrap:wrap}}
.save-bar{{position:sticky;bottom:20px;background:white;border-radius:15px;padding:15px 25px;box-shadow:0 5px 25px rgba(0,0,0,.2);display:flex;align-items:center;justify-content:space-between;margin-top:20px}}
.msg{{padding:10px 20px;border-radius:8px;font-weight:600;display:none}}
.msg-ok{{background:#d4edda;color:#155724}}
.msg-err{{background:#f8d7da;color:#721c24}}
</style></head>
<body>
<nav class="nav">
  <img src="https://upload.wikimedia.org/wikipedia/commons/6/65/Universidad_Tecmilenio.png" alt="Tecmilenio" style="height:36px;filter:brightness(0) invert(1)">
  <div class="nav-links">
    <a href="/">Inicio</a>
    <a href="/admin">← Panel</a>
    <a href="/ver">Resultados</a>
    <a href="/admin/logout">Salir</a>
  </div>
</nav>
<div class="container">
  <h1>✏️ Editor de Examen</h1>

  <!-- Configuración general -->
  <div class="config-card">
    <h3>Configuración General</h3>
    <div class="row">
      <div>
        <label>Título del examen</label>
        <input type="text" id="cfg_titulo" value="{cfg['titulo']}">
      </div>
      <div>
        <label>Materia / Curso</label>
        <input type="text" id="cfg_materia" value="{cfg['materia']}">
      </div>
    </div>
    <div style="margin-top:15px;max-width:250px">
      <label>Tiempo límite (minutos, 0 = sin límite)</label>
      <input type="number" id="cfg_tiempo" value="{cfg['tiempo_minutos']}" min="0">
    </div>
  </div>

  <!-- Barra de herramientas -->
  <div class="toolbar">
    <button class="btn btn-green" onclick="agregarPregunta('opcion_multiple')">+ Opción múltiple</button>
    <button class="btn btn-blue"  onclick="agregarPregunta('seleccion_multiple')">+ Selección múltiple</button>
    <button class="btn btn-orange" onclick="agregarPregunta('abierta')">+ Pregunta abierta</button>
  </div>

  <div id="preguntas-container"></div>

  <div id="msg" class="msg"></div>

  <div class="save-bar">
    <span style="color:#888;font-size:14px">Los cambios se guardan en el servidor</span>
    <button class="btn btn-green" style="font-size:16px;padding:14px 35px" onclick="guardarExamen()">💾 Guardar Examen</button>
  </div>
</div>

<script>
let preguntas = {pregs_json};
let nextId = preguntas.length ? Math.max(...preguntas.map(p=>p.id)) + 1 : 1;

function render() {{
  const c = document.getElementById('preguntas-container');
  c.innerHTML = '';
  preguntas.forEach((p, idx) => c.appendChild(crearCard(p, idx)));
}}

function crearCard(p, idx) {{
  const div = document.createElement('div');
  div.className = 'pregunta-card ' + p.tipo;
  div.id = 'preg_' + p.id;

  const badges = {{opcion_multiple:'Opción múltiple',seleccion_multiple:'Selección múltiple',abierta:'Pregunta abierta'}};
  const colores = {{opcion_multiple:'tipo-opcion_multiple',seleccion_multiple:'tipo-seleccion_multiple',abierta:'tipo-abierta'}};

  let opcionesHTML = '';
  if (p.tipo !== 'abierta') {{
    (p.opciones||[]).forEach((op, oi) => {{
      const checked = p.tipo === 'seleccion_multiple'
        ? (Array.isArray(p.respuesta_correcta) && p.respuesta_correcta.includes(op.id) ? 'checked' : '')
        : (p.respuesta_correcta === op.id ? 'checked' : '');
      const inputType = p.tipo === 'seleccion_multiple' ? 'checkbox' : 'radio';
      opcionesHTML += `
        <div class="opcion-row">
          <span class="opcion-id">${{op.id}}</span>
          <input type="text" value="${{op.texto}}" onchange="editarOpcionTexto(${{p.id}}, ${{oi}}, this.value)" placeholder="Texto de la opción">
          <input type="${{inputType}}" name="correcta_${{p.id}}" class="correcta-check" ${{checked}}
            onchange="marcarCorrecta(${{p.id}}, '${{op.id}}', this.checked, '${{p.tipo}}')"
            title="Marcar como correcta">
          <button class="btn btn-red btn-sm" onclick="eliminarOpcion(${{p.id}}, ${{oi}})">✕</button>
        </div>`;
    }});
    opcionesHTML += `<button class="btn btn-gray btn-sm" style="margin-top:8px" onclick="agregarOpcion(${{p.id}})">+ Opción</button>`;
  }} else {{
    opcionesHTML = '<p style="color:#888;font-size:14px;font-style:italic">Pregunta de respuesta abierta — el alumno escribirá su respuesta.</p>';
  }}

  // Imagen
  const imgSrc = p.imagen ? (p.imagen.startsWith('data:') || p.imagen.startsWith('http') ? p.imagen : '/imagen/' + p.imagen) : '';
  const imgStyle = imgSrc ? 'display:block' : 'display:none';

  div.innerHTML = `
    <div class="pregunta-header">
      <div style="display:flex;align-items:center;gap:12px">
        <div class="pregunta-num">${{idx+1}}</div>
        <span class="tipo-badge ${{colores[p.tipo]}}">${{badges[p.tipo]}}</span>
        <input type="number" value="${{p.puntos||1}}" min="1" max="10" style="width:65px"
          onchange="editarCampo(${{p.id}},'puntos',+this.value)" title="Puntos">
        <label style="font-size:13px;color:#888;margin:0">pts</label>
      </div>
      <div style="display:flex;gap:8px">
        <button class="btn btn-gray btn-sm" onclick="moverPregunta(${{idx}},-1)" title="Subir">↑</button>
        <button class="btn btn-gray btn-sm" onclick="moverPregunta(${{idx}},1)" title="Bajar">↓</button>
        <button class="btn btn-red btn-sm" onclick="eliminarPregunta(${{p.id}})">Eliminar</button>
      </div>
    </div>
    <div style="margin-bottom:12px">
      <label>Pregunta</label>
      <textarea rows="2" style="resize:vertical" onchange="editarCampo(${{p.id}},'texto',this.value)">${{p.texto||''}}</textarea>
    </div>
    <div style="margin-bottom:15px">
      <label>Imagen (opcional)</label>
      <input type="file" accept="image/*" onchange="cargarImagen(${{p.id}}, this)">
      <img id="img_${{p.id}}" src="${{imgSrc}}" class="img-preview" style="${{imgStyle}}">
      ${{imgSrc ? '<button class="btn btn-red btn-sm" style="margin-top:5px" onclick="quitarImagen('+p.id+')">✕ Quitar imagen</button>' : ''}}
    </div>
    <div>
      ${{p.tipo !== 'abierta' ? '<label>Opciones <small style="color:#888">(marca la correcta)</small></label>' : ''}}
      ${{opcionesHTML}}
    </div>`;
  return div;
}}

function editarCampo(id, campo, val) {{
  const p = preguntas.find(x=>x.id===id);
  if (p) p[campo] = val;
}}
function editarOpcionTexto(id, oi, val) {{
  const p = preguntas.find(x=>x.id===id);
  if (p && p.opciones) p.opciones[oi].texto = val;
}}
function marcarCorrecta(id, opId, checked, tipo) {{
  const p = preguntas.find(x=>x.id===id);
  if (!p) return;
  if (tipo === 'seleccion_multiple') {{
    if (!Array.isArray(p.respuesta_correcta)) p.respuesta_correcta = [];
    if (checked) {{ if (!p.respuesta_correcta.includes(opId)) p.respuesta_correcta.push(opId); }}
    else p.respuesta_correcta = p.respuesta_correcta.filter(x=>x!==opId);
  }} else {{
    p.respuesta_correcta = opId;
  }}
}}
function agregarOpcion(id) {{
  const p = preguntas.find(x=>x.id===id);
  if (!p) return;
  const ids = ['a','b','c','d','e','f','g','h'];
  const nextOp = ids[p.opciones.length] || String.fromCharCode(97+p.opciones.length);
  p.opciones.push({{id: nextOp, texto: ''}});
  render();
}}
function eliminarOpcion(id, oi) {{
  const p = preguntas.find(x=>x.id===id);
  if (p) p.opciones.splice(oi, 1);
  render();
}}
function eliminarPregunta(id) {{
  if (!confirm('¿Eliminar esta pregunta?')) return;
  preguntas = preguntas.filter(x=>x.id!==id);
  render();
}}
function moverPregunta(idx, dir) {{
  const ni = idx + dir;
  if (ni < 0 || ni >= preguntas.length) return;
  [preguntas[idx], preguntas[ni]] = [preguntas[ni], preguntas[idx]];
  render();
}}
function agregarPregunta(tipo) {{
  const base = {{id: nextId++, tipo, texto:'', imagen:'', puntos:1}};
  if (tipo !== 'abierta') {{
    base.opciones = [
      {{id:'a', texto:''}}, {{id:'b', texto:''}},
      {{id:'c', texto:''}}, {{id:'d', texto:''}}
    ];
    base.respuesta_correcta = tipo === 'seleccion_multiple' ? [] : 'a';
  }}
  preguntas.push(base);
  render();
  setTimeout(()=>document.getElementById('preg_'+base.id)?.scrollIntoView({{behavior:'smooth'}}), 100);
}}
function cargarImagen(id, input) {{
  const file = input.files[0];
  if (!file) return;
  const reader = new FileReader();
  reader.onload = e => {{
    const p = preguntas.find(x=>x.id===id);
    if (p) p.imagen = e.target.result;
    const img = document.getElementById('img_'+id);
    if (img) {{ img.src = e.target.result; img.style.display='block'; }}
  }};
  reader.readAsDataURL(file);
}}
function quitarImagen(id) {{
  const p = preguntas.find(x=>x.id===id);
  if (p) p.imagen = '';
  render();
}}
function showMsg(txt, ok) {{
  const el = document.getElementById('msg');
  el.textContent = txt;
  el.className = 'msg ' + (ok ? 'msg-ok' : 'msg-err');
  el.style.display = 'block';
  setTimeout(()=>el.style.display='none', 4000);
}}
async function guardarExamen() {{
  // Actualizar textos de textareas (onchange no siempre dispara)
  preguntas.forEach(p => {{
    const card = document.getElementById('preg_'+p.id);
    if (card) {{
      const ta = card.querySelector('textarea');
      if (ta) p.texto = ta.value;
    }}
  }});
  const cfg = {{
    titulo: document.getElementById('cfg_titulo').value,
    materia: document.getElementById('cfg_materia').value,
    tiempo_minutos: parseInt(document.getElementById('cfg_tiempo').value)||0,
    preguntas
  }};
  try {{
    const r = await fetch('/admin/guardar_config', {{
      method:'POST', headers:{{'Content-Type':'application/json'}},
      credentials:'include', body: JSON.stringify(cfg)
    }});
    const d = await r.json();
    showMsg(d.ok ? 'Examen guardado correctamente' : 'Error: ' + d.error, d.ok);
  }} catch(e) {{
    showMsg('Error al guardar: ' + e.message, false);
  }}
}}
render();

// Avisar al servidor que el docente esta editando
async function notificarEditando(estado) {{
    try {{
        await fetch('/admin/set_editando', {{
            method: 'POST',
            headers: {{'Content-Type': 'application/json'}},
            credentials: 'include',
            body: JSON.stringify({{editando: estado}})
        }});
    }} catch(e) {{}}
}}

// Activar modo edicion al primer cambio
let modoEdicionActivo = false;
document.addEventListener('input', function() {{
    if (!modoEdicionActivo) {{
        modoEdicionActivo = true;
        notificarEditando(true);
    }}
}}, true);

// Desactivar si cierra la pestana sin guardar
window.addEventListener('beforeunload', function() {{
    notificarEditando(false);
}});
</script>
</body></html>'''

@app.route('/admin/set_editando', methods=['POST'])
@admin_required
def set_editando():
    data = request.get_json(silent=True) or {}
    with _estado_lock:
        _estado_edicion['editando'] = data.get('editando', False)
    return jsonify({'ok': True})

@app.route('/admin/guardar_config', methods=['POST'])
@admin_required
def admin_guardar_config():
    try:
        cfg = request.json
        guardar_config(cfg)
        with _estado_lock:
            _estado_edicion['editando'] = False
            _estado_edicion['version'] += 1
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# ── Descarga Excel ────────────────────────────────────────────

@app.route('/admin/descargar')
@admin_required
def admin_descargar():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return '<h2>Instala openpyxl: pip install openpyxl</h2>', 500

    archivos = sorted(os.listdir('respuestas'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Resultados'

    verde   = PatternFill('solid', fgColor='D4EDDA')
    naranja = PatternFill('solid', fgColor='FFE8C8')
    rojo    = PatternFill('solid', fgColor='F8D7DA')
    header_fill = PatternFill('solid', fgColor='00A859')
    header_font = Font(color='FFFFFF', bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Nombre', 'Calificación', 'Correctas', 'Total Puntos', 'Estado', 'Fecha y Hora']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for ri, archivo in enumerate(archivos, 2):
        try:
            with open(f'respuestas/{archivo}', 'r', encoding='utf-8') as f:
                d = json.load(f)
        except:
            continue
        cal   = float(d.get('calificacion', 0))
        estado = 'Aprobado' if cal >= 6 else 'Reprobado'
        fill   = verde if cal >= 8 else (naranja if cal >= 6 else rojo)
        row    = [d.get('nombre',''), cal, d.get('correctas','?'),
                  d.get('total_puntos','?'), estado, d.get('timestamp','')]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = center
            cell.border = thin
            if ci in (2, 5):
                cell.fill = fill
                if ci == 2:
                    cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'resultados_examen_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ══════════════════════════════════════════════════════════════
#  PÁGINA PÚBLICA DE RESULTADOS
# ══════════════════════════════════════════════════════════════

@app.route('/')
def home():
    total    = len(os.listdir('respuestas'))
    logo_svg = 'https://upload.wikimedia.org/wikipedia/commons/6/65/Universidad_Tecmilenio.png'
    return f'''<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<title>Servidor de Exámenes | Tecmilenio</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:linear-gradient(135deg,#00A859,#007A3D);min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px}}
.card{{background:white;border-radius:20px;padding:40px 50px;text-align:center;box-shadow:0 20px 60px rgba(0,0,0,.3);max-width:500px;width:100%;border-top:6px solid #00A859}}
.logo{{margin-bottom:20px}} .logo img{{max-width:200px}}
.status{{display:inline-flex;align-items:center;gap:8px;background:#d4edda;padding:10px 20px;border-radius:50px;margin-bottom:20px}}
.dot{{width:12px;height:12px;background:#28a745;border-radius:50%;animation:pulse 2s infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.5}}}}
.status-text{{color:#155724;font-weight:700;font-size:14px}}
h1{{color:#007A3D;font-size:26px;margin-bottom:5px}}
.sub{{color:#888;font-size:14px;margin-bottom:25px}}
.count{{font-size:68px;font-weight:900;color:#00A859;line-height:1}}
.lbl{{color:#007A3D;font-size:15px;font-weight:600;text-transform:uppercase;letter-spacing:1px;margin-bottom:25px}}
.btn{{display:inline-block;background:linear-gradient(135deg,#00A859,#007A3D);color:white;text-decoration:none;padding:15px 40px;border-radius:50px;font-size:17px;font-weight:700;box-shadow:0 5px 20px rgba(0,168,89,.4);transition:.2s;margin:5px}}
.btn:hover{{transform:translateY(-2px)}}
.btn-ghost{{background:transparent;border:2px solid #00A859;color:#00A859;box-shadow:none}}
.footer{{margin-top:25px;padding-top:20px;border-top:1px solid #eee;color:#aaa;font-size:13px}}
</style></head>
<body><div class="card">
  <div class="logo"><div style="background:white;border-radius:10px;padding:8px 18px;display:inline-block"><img src="{logo_svg}" alt="Tecmilenio" style="max-height:50px"></div></div>
  <div class="status"><span class="dot"></span><span class="status-text">SERVIDOR ACTIVO</span></div>
  <h1>Sistema de Exámenes</h1>
  <p class="sub">Gestión de Redes | Tecmilenio 2026</p>
  <div class="count">{total}</div>
  <div class="lbl">{"Respuesta recibida" if total==1 else "Respuestas recibidas"}</div>
  <a href="/ver" class="btn">Ver Resultados</a>
  <a href="/admin" class="btn btn-ghost">Panel Docente</a>
  <div class="footer">Tecmilenio &nbsp;|&nbsp; Sistema desarrollado para Gestión de Redes</div>
</div></body></html>'''

@app.route('/ver')
def ver_respuestas():
    archivos = sorted(os.listdir('respuestas'))
    logo_svg = 'https://upload.wikimedia.org/wikipedia/commons/6/65/Universidad_Tecmilenio.png'

    if not archivos:
        return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8"><title>Sin resultados</title>
<style>body{{font-family:Arial;background:linear-gradient(135deg,#00A859,#007A3D);min-height:100vh;display:flex;align-items:center;justify-content:center}}
.card{{background:white;border-radius:20px;padding:60px;text-align:center;box-shadow:0 10px 40px rgba(0,0,0,.3);border-top:6px solid #00A859}}
img{{width:180px;margin-bottom:25px}}h2{{color:#888}}p{{color:#aaa;margin:15px 0}}a{{color:#00A859;font-weight:700;text-decoration:none}}</style></head>
<body><div class="card"><div style="background:white;border-radius:8px;padding:5px 12px;display:inline-flex;align-items:center"><img src="{logo_svg}" style="height:38px"></div><div style="font-size:48px;opacity:.15">[ ]</div>
<h2>No hay respuestas aún</h2><p>Esperando que los alumnos envíen su examen...</p><a href="/">← Volver</a></div></body></html>'''

    calificaciones = []
    filas = ''
    for archivo in archivos:
        try:
            with open(f'respuestas/{archivo}', 'r', encoding='utf-8') as f:
                d = json.load(f)
        except:
            continue
        nombre    = d.get('nombre', 'Sin nombre')
        fecha     = d.get('timestamp', 'Sin fecha')
        cal       = d.get('calificacion', None)
        correctas = d.get('correctas', '?')
        total     = d.get('total_puntos', 10)
        if cal is not None:
            cal = float(cal)
            calificaciones.append(cal)
            color, estado, bg = ('#28a745','Aprobado','#d4edda') if cal>=8 else \
                                (('#fd7e14','Regular','#fff3cd') if cal>=6 else ('#dc3545','Reprobado','#f8d7da'))
            cal_html = f'<div style="display:inline-flex;align-items:center;gap:8px"><span style="color:{color};font-weight:900;font-size:24px">{cal}</span><span style="background:{bg};color:{color};padding:3px 10px;border-radius:20px;font-size:12px;font-weight:700">{estado}</span></div>'
        else:
            cal_html = '<span style="color:#aaa;font-style:italic">Sin calificación</span>'
        filas += f'<tr><td style="padding:14px 20px;border-bottom:1px solid #f0f0f0;font-weight:600">{nombre}</td><td style="padding:14px 20px;border-bottom:1px solid #f0f0f0;text-align:center">{cal_html}</td><td style="padding:14px 20px;border-bottom:1px solid #f0f0f0;text-align:center;font-weight:600">{correctas}<span style="color:#aaa">/{total}</span></td><td style="padding:14px 20px;border-bottom:1px solid #f0f0f0;color:#aaa;font-size:13px">{fecha}</td></tr>'

    promedio  = round(sum(calificaciones)/len(calificaciones),1) if calificaciones else 0
    cp        = '#28a745' if promedio>=8 else ('#fd7e14' if promedio>=6 else '#dc3545')
    aprobados = sum(1 for c in calificaciones if c>=6)
    reprobados= len(calificaciones)-aprobados
    pct       = round(aprobados/len(calificaciones)*100) if calificaciones else 0

    return f'''<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta http-equiv="refresh" content="15">
<title>Resultados | Tecmilenio</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,sans-serif;background:linear-gradient(135deg,#e8f5e9,#f0f9f4);min-height:100vh;padding:25px 20px}}
.container{{max-width:1100px;margin:0 auto}}
.header{{background:linear-gradient(135deg,#00A859,#007A3D);color:white;padding:25px 35px;border-radius:15px;margin-bottom:25px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:15px;box-shadow:0 5px 20px rgba(0,168,89,.3)}}
.header img{{height:45px}}
.header h1{{font-size:28px;margin:0}}
.badge{{background:rgba(255,255,255,.2);padding:6px 16px;border-radius:20px;font-size:12px;font-weight:700;letter-spacing:1px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:18px;margin-bottom:25px}}
.stat{{background:white;border-radius:12px;padding:22px;text-align:center;box-shadow:0 3px 12px rgba(0,0,0,.08);border-top:5px solid #00A859;transition:.2s}}
.stat:hover{{transform:translateY(-3px)}}
.stat.v{{border-top-color:#28a745}}.stat.r{{border-top-color:#dc3545}}.stat.o{{border-top-color:#fd7e14}}
.snum{{font-size:48px;font-weight:900;line-height:1;margin-bottom:5px}}
.slbl{{color:#888;font-size:12px;text-transform:uppercase;letter-spacing:.5px}}
.barra-wrap{{background:white;border-radius:12px;padding:20px 25px;margin-bottom:25px;box-shadow:0 3px 12px rgba(0,0,0,.08)}}
.barra-track{{background:#e9ecef;border-radius:10px;height:14px;overflow:hidden}}
.barra-fill{{height:100%;background:linear-gradient(90deg,#00A859,#28a745);border-radius:10px;width:{pct}%}}
.table-wrap{{background:white;border-radius:12px;box-shadow:0 3px 12px rgba(0,0,0,.08);overflow:hidden;margin-bottom:25px}}
table{{width:100%;border-collapse:collapse}}
th{{background:linear-gradient(135deg,#00A859,#007A3D);color:white;padding:16px 20px;text-align:left;font-size:13px;text-transform:uppercase;letter-spacing:.5px}}
tr:hover td{{background:#f0f9f4}}
.footer{{display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:10px;background:white;border-radius:12px;padding:15px 25px;box-shadow:0 3px 12px rgba(0,0,0,.08)}}
a{{color:#00A859;text-decoration:none;font-weight:700}}a:hover{{text-decoration:underline}}
.refresh{{color:#aaa;font-size:13px}}
</style></head>
<body><div class="container">
<div class="header">
  <div style="display:flex;align-items:center;gap:20px"><div style="background:white;border-radius:8px;padding:5px 12px;display:inline-flex;align-items:center"><img src="{logo_svg}" style="height:38px"></div>
  <div><h1>Resultados del Examen</h1><span class="badge">GESTIÓN DE REDES 2026</span></div></div>
  <a href="/admin/descargar" style="background:rgba(255,255,255,.15);color:white;padding:10px 20px;border-radius:25px;font-weight:700;text-decoration:none">📥 Descargar Excel</a>
</div>
<div class="grid">
  <div class="stat"><div class="snum" style="color:#00A859">{len(archivos)}</div><div class="slbl">Alumnos</div></div>
  <div class="stat"><div class="snum" style="color:{cp}">{promedio}</div><div class="slbl">Promedio</div></div>
  <div class="stat v"><div class="snum" style="color:#28a745">{aprobados}</div><div class="slbl">Aprobados</div></div>
  <div class="stat r"><div class="snum" style="color:#dc3545">{reprobados}</div><div class="slbl">Reprobados</div></div>
  <div class="stat o"><div class="snum" style="color:#fd7e14">{pct}%</div><div class="slbl">% Aprobacion</div></div>
</div>
<div class="barra-wrap">
  <div style="font-size:13px;color:#888;font-weight:600;text-transform:uppercase;letter-spacing:.5px;margin-bottom:10px">Tasa de aprobación</div>
  <div class="barra-track"><div class="barra-fill"></div></div>
</div>
<div class="table-wrap"><table>
<tr><th>Nombre</th><th style="text-align:center">Calificacion</th><th style="text-align:center">Correctas</th><th>Fecha y Hora</th></tr>
{filas}
</table></div>
<div class="footer">
  <div style="display:flex;gap:20px;align-items:center">
    <a href="/">← Inicio</a>
    <span style="color:#ddd">|</span>
    <a href="/admin">Panel Docente</a>
  </div>
  <span class="refresh">Se actualiza cada 15 segundos</span>
</div>
</div></body></html>'''

# ══════════════════════════════════════════════════════════════
#  ARRANQUE
# ══════════════════════════════════════════════════════════════

LOGO_TEC = """
╔══════════════════════════════════════════════════════════════╗
║                                                              ║
║         TECMILENIO — SISTEMA DE EXÁMENES v2.0                ║
║         Gestión de Redes | 2026                              ║
║                                                              ║
╚══════════════════════════════════════════════════════════════╝
"""

if __name__ == '__main__':
    ips = get_ips()
    print('\033[92m' + LOGO_TEC + '\033[0m')
    print('Servidor iniciado correctamente\n')
    print('Direcciones disponibles:\n')
    for ip in ips:
        print(f'   -- http://{ip}:5000         \033[90m← Página principal\033[0m')
        print(f'   -- http://{ip}:5000/ver     \033[90m← Resultados públicos\033[0m')
        print(f'   -- http://{ip}:5000/admin   \033[90m← Panel docente (contraseña: {ADMIN_PASS})\033[0m')
        print()
    print('INFO: Instala openpyxl para descarga Excel:  pip install openpyxl')
    print('Presiona Ctrl+C para detener\n')
    app.run(host='0.0.0.0', port=5000, debug=False)
